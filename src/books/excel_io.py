"""Excel I/O helpers: export, template write and read.

Separated from `books.excel` to reduce monolithic file size.
"""
import logging
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - openpyxl ships with the app
    openpyxl = None

from books import tags
from books.model import Book, is_valid_isbn, normalize_isbn
from books.reading import parse_stamp, reading_days, status_of
from shared.texts import field_label, text

logger = logging.getLogger("katipcelebi")

TEMPLATE_DEFAULT_NAME = "isbn_list.xlsx"
EXPORT_DEFAULT_NAME = "my_library.xlsx"

MAX_IMPORT_BYTES = 20 * 1024 * 1024  # 20 MiB
_FORMULA_START = ("=", "+", "-", "@")


def _formula_safe(value):
    if isinstance(value, str) and value.startswith(_FORMULA_START):
        return "'" + value
    return value


DATE_FIELDS = ("started_date", "finished_date")

EXPORTED_FIELDS = (
    "title",
    "subtitle",
    "authors",
    "publishers",
    "publish_date",
    "publish_places",
    "edition_name",
    "series",
    "number_of_pages",
    "languages",
    "isbn_10",
    "isbn_13",
    "subjects",
    "rating",
    "status",
    "tags",
    "signed",
    "copies",
    "notes",
    "started_date",
    "finished_date",
)

DAYS_COLUMN = "days"
LENT_COLUMN = "lent_to"
COLUMN_NAMES = ("isbn",) + EXPORTED_FIELDS + (DAYS_COLUMN, LENT_COLUMN)


def _stamp(value: str):
    return parse_stamp(value) or value or ""


def export_library(books: list[Book], ledger, path: Path) -> bool:
    if openpyxl is None:
        logger.error("openpyxl is not installed; cannot export")
        return False

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = text("nav_library")[:31]

    headings = [field_label("isbn")] + [field_label(name) for name in EXPORTED_FIELDS]
    headings += [text("export_days"), text("export_lent_to")]
    for column, heading in enumerate(headings, start=1):
        cell = sheet.cell(row=1, column=column, value=heading)
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"

    for row, entry in enumerate(books, start=2):
        values = [entry.display_isbn]
        for name in EXPORTED_FIELDS:
            values.append(_column_value(entry, name))
        days = reading_days(entry)
        values.append(round(days, 2) if days else "")
        out = ledger.open_loans_for(entry.key) if ledger is not None else []
        values.append(", ".join(loan.person_name for loan in out))

        for column, value in enumerate(values, start=1):
            safe = _formula_safe(value)
            cell = sheet.cell(row=row, column=column, value=safe)
            if safe is not value and hasattr(cell, "quotePrefix"):
                cell.quotePrefix = True
            name = COLUMN_NAMES[column - 1]
            if name in ("isbn", "isbn_10", "isbn_13"):
                cell.number_format = "@"
            elif name == DAYS_COLUMN and value != "":
                cell.number_format = "0.0"
            elif name in DATE_FIELDS and value:
                cell.number_format = "yyyy-mm-dd hh:mm"

    _widen(sheet, headings)
    try:
        book.save(path)
        logger.info("Exported %d book(s) to %s", len(books), path)
        return True
    except (OSError, ValueError):
        logger.exception("Could not write %s", path)
        return False


def _column_value(entry: Book, name: str):
    if name in DATE_FIELDS:
        return _stamp(getattr(entry, name))
    if name == "status":
        return text("status_" + status_of(entry))
    if name == "tags":
        return tags.show(entry.tags)
    if name == "signed":
        return text("export_yes") if entry.signed.strip() else text("export_no")
    return getattr(entry, name)


def _widen(sheet, headings: list[str]) -> None:
    for column, heading in enumerate(headings, start=1):
        longest = len(str(heading))
        for cell in sheet[get_column_letter(column)]:
            longest = max(longest, len(str(cell.value or "")))
        sheet.column_dimensions[get_column_letter(column)].width = min(
            48, max(10, longest + 2)
        )


def write_template(path: Path) -> bool:
    if openpyxl is None:
        return False
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = field_label("isbn")
    cell = sheet.cell(row=1, column=1, value=field_label("isbn"))
    cell.font = Font(bold=True)
    sheet.column_dimensions["A"].width = 22
    for row in range(2, 500):
        sheet.cell(row=row, column=1).number_format = "@"
    try:
        book.save(path)
        return True
    except (OSError, ValueError):
        logger.exception("Could not write %s", path)
        return False


class FileTooLarge(Exception):
    pass


def read_template(path: Path) -> list[str] | None:
    if openpyxl is None:
        return None
    try:
        size = path.stat().st_size
    except OSError:
        logger.exception("Could not inspect %s", path)
        return None
    if size > MAX_IMPORT_BYTES:
        logger.warning("Refusing to read %s: %d bytes", path, size)
        raise FileTooLarge(path)
    try:
        book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        logger.exception("Could not read %s", path)
        return None

    found, seen = [], set()
    try:
        for sheet in book.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for value in row:
                    if value is None:
                        continue
                    raw = (
                        str(int(value))
                        if isinstance(value, float) and value.is_integer()
                        else str(value)
                    )
                    isbn = normalize_isbn(raw)
                    if is_valid_isbn(isbn) and isbn not in seen:
                        seen.add(isbn)
                        found.append(isbn)
    finally:
        book.close()
    return found
